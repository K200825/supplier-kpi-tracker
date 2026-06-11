"""
export_report.py
----------------
Takes the scored DataFrame and exports a formatted Excel report with:
  - Sheet 1: KPI Scorecard (colour-coded tiers, score bars)
  - Sheet 2: Raw Data (source data for audit trail)
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ── Colour palette ────────────────────────────────────────────────────────────
COLOURS = {
    "header_bg":    "1F3864",   # dark navy
    "header_font":  "FFFFFF",   # white
    "tier_a":       "C6EFCE",   # green fill
    "tier_a_font":  "276221",
    "tier_b":       "FFEB9C",   # amber fill
    "tier_b_font":  "9C6500",
    "tier_c":       "FFC7CE",   # red fill
    "tier_c_font":  "9C0006",
    "alt_row":      "F2F2F2",   # light grey alternate row
    "border":       "BFBFBF",
}

TIER_COLOURS = {
    "A – Strategic": (COLOURS["tier_a"],      COLOURS["tier_a_font"]),
    "B – Approved":  (COLOURS["tier_b"],      COLOURS["tier_b_font"]),
    "C – At Risk":   (COLOURS["tier_c"],      COLOURS["tier_c_font"]),
}


def _thin_border():
    side = Side(style="thin", color=COLOURS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _header_cell(cell, text):
    cell.value = text
    cell.font      = Font(name="Arial", bold=True, color=COLOURS["header_font"], size=10)
    cell.fill      = PatternFill("solid", start_color=COLOURS["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()


def _data_cell(cell, value, row_idx, fmt=None, align="center"):
    cell.value = value
    bg = COLOURS["alt_row"] if row_idx % 2 == 0 else "FFFFFF"
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _thin_border()
    if fmt:
        cell.number_format = fmt


def build_scorecard_sheet(ws, df: pd.DataFrame):
    """Sheet 1 – KPI Scorecard"""
    ws.title = "KPI Scorecard"
    ws.sheet_view.showGridLines = False

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title = ws["A1"]
    title.value     = "Supplier KPI Scorecard"
    title.font      = Font(name="Arial", bold=True, size=14, color=COLOURS["header_bg"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Subtitle / date ───────────────────────────────────────────────────────
    ws.merge_cells("A2:J2")
    subtitle = ws["A2"]
    from datetime import date
    subtitle.value = f"Generated: {date.today().strftime('%d %B %Y')}  |  Weights: OTD 45% | Lead Time 35% | Defect Rate 20%"
    subtitle.font  = Font(name="Arial", italic=True, size=9, color="595959")
    subtitle.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        "Rank", "Supplier ID", "Supplier Name",
        "OTD Rate", "OTD Score\n(0-100)",
        "Lead Time\nRatio", "LTC Score\n(0-100)",
        "Defect Rate", "Defect Score\n(0-100)",
        "Composite\nScore", "Tier"
    ]
    ws.row_dimensions[3].height = 36
    for col_idx, h in enumerate(headers, start=1):
        _header_cell(ws.cell(row=3, column=col_idx), h)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, row in df.iterrows():
        r = i + 4  # Excel row (1-indexed, offset by header rows)
        ws.row_dimensions[r].height = 18

        cells_vals = [
            (row["rank"],                       None,    "center"),
            (row["supplier_id"],                None,    "center"),
            (row["supplier_name"],              None,    "left"),
            (row["otd_rate"],                   "0.0%",  "center"),
            (round(row["otd_score"], 1),        "0.0",   "center"),
            (row["lead_time_ratio"],            "0.00x", "center"),
            (round(row["ltc_score"], 1),        "0.0",   "center"),
            (row["defect_rate"],                "0.000%","center"),
            (round(row["defect_score"], 1),     "0.0",   "center"),
            (row["composite_score"],            "0.0",   "center"),
            (row["tier"],                       None,    "center"),
        ]

        for col_idx, (val, fmt, align) in enumerate(cells_vals, start=1):
            cell = ws.cell(row=r, column=col_idx)
            _data_cell(cell, val, i, fmt=fmt, align=align)

        # Colour-code the Tier cell and the whole row lightly
        tier_cell = ws.cell(row=r, column=11)
        bg, fg = TIER_COLOURS.get(row["tier"], ("FFFFFF", "000000"))
        tier_cell.fill = PatternFill("solid", start_color=bg)
        tier_cell.font = Font(name="Arial", bold=True, size=10, color=fg)

    # ── Conditional formatting: colour scale on composite score ──────────────
    last_row = 3 + len(df)
    score_col = "J"
    ws.conditional_formatting.add(
        f"{score_col}4:{score_col}{last_row}",
        ColorScaleRule(
            start_type="min",  start_color="FFC7CE",
            mid_type="percentile", mid_value=50, mid_color="FFEB9C",
            end_type="max",    end_color="C6EFCE",
        )
    )

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [6, 12, 28, 10, 12, 12, 12, 12, 14, 12, 16]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ── Freeze panes below header ─────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Summary box bottom ────────────────────────────────────────────────────
    summary_row = last_row + 2
    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    ws.cell(row=summary_row, column=1).value = "Tier Summary"
    ws.cell(row=summary_row, column=1).font  = Font(name="Arial", bold=True, size=10)

    tier_counts = df["tier"].value_counts()
    for offset, tier in enumerate(["A – Strategic", "B – Approved", "C – At Risk"], start=1):
        label_cell = ws.cell(row=summary_row + offset, column=1)
        count_cell = ws.cell(row=summary_row + offset, column=2)
        label_cell.value = tier
        count_cell.value = int(tier_counts.get(tier, 0))
        bg, fg = TIER_COLOURS.get(tier, ("FFFFFF", "000000"))
        label_cell.fill = PatternFill("solid", start_color=bg)
        label_cell.font = Font(name="Arial", bold=True, color=fg, size=10)
        count_cell.font = Font(name="Arial", size=10)


def build_raw_sheet(ws, df: pd.DataFrame):
    """Sheet 2 – Raw Source Data (audit trail)"""
    ws.title = "Raw Data"
    ws.sheet_view.showGridLines = False

    raw_cols = [
        "supplier_id", "supplier_name", "orders_total", "orders_on_time",
        "promised_lead_time_days", "avg_lead_time_days",
        "units_delivered", "defects_total"
    ]
    headers = [
        "Supplier ID", "Supplier Name", "Total Orders", "On-Time Orders",
        "Promised Lead Time (days)", "Actual Lead Time (days)",
        "Units Delivered", "Total Defects"
    ]

    ws.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, start=1):
        _header_cell(ws.cell(row=1, column=col_idx), h)

    for i, row in df.iterrows():
        r = i + 2
        ws.row_dimensions[r].height = 16
        for col_idx, col in enumerate(raw_cols, start=1):
            align = "left" if col == "supplier_name" else "center"
            _data_cell(ws.cell(row=r, column=col_idx), row[col], i, align=align)

    col_widths = [12, 28, 14, 14, 24, 22, 16, 14]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"


def export_report(df: pd.DataFrame, output_path: str):
    wb = Workbook()
    ws1 = wb.active
    build_scorecard_sheet(ws1, df)

    ws2 = wb.create_sheet()
    build_raw_sheet(ws2, df)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Report saved → {output_path}")
    return output_path


if __name__ == "__main__":
    import sys, os
    sys.path.insert(0, os.path.dirname(__file__))
    from generate_data import generate_supplier_data
    from kpi_scoring import run_scoring

    raw = generate_supplier_data()
    scored = run_scoring(raw)
    out = os.path.join(os.path.dirname(__file__), "..", "output", "supplier_kpi_report.xlsx")
    export_report(scored, out)
